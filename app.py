from __future__ import annotations

import base64
import io
import ipaddress
import re
import urllib.request
import zipfile
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from supabase import Client, create_client

from parsers import parse_pdf


APP_VERSION = "Snoopy IA"
AUTHOR = "@PamperoSur"
AUTHOR_CREDIT = "X: @PamperoSur · CAF"
TZ_AR = ZoneInfo("America/Argentina/Buenos_Aires")
MONTHS = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
          "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
DISCLAIMER = (
    "Aplicación académica desarrollada en el marco de Emprendedurismo (IA) - Corrientes. "
    "Uso exclusivo de estudiantes y docentes autorizados. No constituye asesoramiento ni "
    "prestación de servicios contables, impositivos, financieros, jurídicos o profesionales. "
    "Los resultados deben verificarse contra la documentación bancaria original. Se prohíbe "
    "su uso o explotación comercial. El autor no asume responsabilidad por errores de "
    "extracción, interpretación o decisiones tomadas con la información procesada."
)

st.set_page_config(page_title="Snoopy IA | PDF bancario → Excel", page_icon="🏦", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=IBM+Plex+Mono:wght@500;600&display=swap');
:root { --navy:#061525; --panel:#0a2139; --cyan:#48d7d0; --blue:#1597d5; --neon:#39f2a0; --neon-dark:#0aae77; --line:#245573; --muted:#9eb4c8; }
html, body, [class*="css"] { font-family:Inter,sans-serif; }
[data-testid="stAppViewContainer"] { background:
 linear-gradient(rgba(34,86,119,.055) 1px,transparent 1px),linear-gradient(90deg,rgba(34,86,119,.055) 1px,transparent 1px),
 radial-gradient(circle at 88% 3%,rgba(12,152,178,.18),transparent 25%),linear-gradient(145deg,#051321,#09213a 58%,#061725);
 background-size:38px 38px,38px 38px,auto,auto; color:#eef6ff; }
[data-testid="stSidebar"] { background:#061321; border-right:1px solid #1d4463; }
[data-testid="stHeader"], [data-testid="stToolbar"], [data-testid="stAppToolbar"],
[data-testid="stDecoration"], [data-testid="stStatusWidget"], #MainMenu, footer {
 display:none !important; visibility:hidden !important; height:0 !important; }
.block-container { max-width:1500px; padding-top:0 !important; padding-bottom:.2rem !important; margin-top:-.45rem; }
.hero { display:grid; grid-template-columns:minmax(0,1fr) auto; gap:2rem; align-items:stretch; padding:1.6rem 1.75rem;
 border-radius:20px; border:1px solid #28617f; background:radial-gradient(circle at 83% 18%,rgba(24,184,190,.26),transparent 34%),linear-gradient(135deg,rgba(12,44,72,.98),rgba(7,27,47,.98));
 box-shadow:0 22px 60px rgba(0,0,0,.28),inset 0 1px 0 rgba(255,255,255,.05); margin-bottom:.55rem; overflow:hidden; }
.hero h1 { margin:.28rem 0 .5rem; color:#9fb8ca; font-size:clamp(1.75rem,3vw,2.45rem); letter-spacing:-.035em;
 font-weight:720; text-shadow:0 2px 10px rgba(0,0,0,.32); align-self:end; }
.title-arrow { color:#50ddd4; text-shadow:0 0 14px rgba(80,221,212,.2); }
.hero p { margin:0; color:#b7cadb; max-width:900px; }
.eyebrow { color:#50d3c8; font-size:.76rem; font-weight:800; letter-spacing:.14em; text-transform:uppercase; }
.brand-lockup { min-width:285px; align-self:stretch; display:flex; flex-direction:column; justify-content:flex-end; padding:0 .15rem .1rem 0; text-align:right; background:transparent; border:0; }
.brand-name { font:800 clamp(1.85rem,2.55vw,2.55rem) 'IBM Plex Mono',monospace; letter-spacing:.055em; line-height:1; white-space:nowrap;
 background:linear-gradient(105deg,#8ff8ef 0%,#48d7d0 38%,#55bfff 82%); -webkit-background-clip:text; background-clip:text; color:transparent;
 filter:drop-shadow(0 3px 10px rgba(0,0,0,.78)) drop-shadow(0 0 16px rgba(50,210,216,.18)); }
.brand-ia { margin-left:.22em; }
.brand-author { margin-top:.58rem; color:#eaf8ff; font:600 .74rem 'IBM Plex Mono',monospace; letter-spacing:.045em;
 text-shadow:0 2px 8px rgba(0,0,0,.9); display:flex; align-items:center; justify-content:flex-end; gap:.48rem; }
.brand-xicon { display:inline-grid; place-items:center; width:1.55rem; height:1.55rem; border:1px solid #46d9e8; border-radius:50%;
 color:#fff; font-size:.9rem; line-height:1; box-shadow:0 0 14px rgba(70,217,232,.16); }
.hero-copy { min-width:0; display:grid; grid-template-rows:auto 1fr auto; }
.hero-copy p { margin-top:0; padding-top:.72rem; }
.hero-meta { margin-top:.48rem; color:#78ddd7; text-align:right;
 font:600 .64rem 'IBM Plex Mono',monospace; letter-spacing:.11em; text-transform:uppercase; text-shadow:0 2px 8px rgba(0,0,0,.75); }
.badge { display:inline-block; margin:.9rem .4rem 0 0; padding:.3rem .62rem; border-radius:999px; font-size:.72rem;
 background:#123451; border:1px solid #2b688e; color:#e3f3ff; }
.legal-title { color:#f2c65c; font:600 .72rem 'IBM Plex Mono',monospace; letter-spacing:.12em; text-transform:uppercase; margin-bottom:.45rem; }
.legal { margin-top:1.5rem; padding:1rem 1.2rem; border-radius:12px; border-left:4px solid #d6aa47;
 background:#13263c; color:#b7c7d8; font-size:.78rem; line-height:1.55; }
.author { color:#fff; font-weight:700; }
.status-grid { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:.68rem; margin:.18rem 0 .45rem; }
.status-card { position:relative; padding:.68rem 1rem; border-radius:14px; border:1px solid rgba(52,110,144,.7);
 background:linear-gradient(145deg,rgba(13,43,68,.92),rgba(7,29,49,.92)); overflow:hidden; }
.status-card:before { content:""; position:absolute; left:0; top:0; bottom:0; width:3px; background:linear-gradient(#47ddd2,#1597d5); }
.status-label { color:#8faabd; font:600 .64rem 'IBM Plex Mono',monospace; letter-spacing:.1em; text-transform:uppercase; }
.status-value { margin-top:.28rem; color:#f7fbff; font-size:1.08rem; font-weight:750; }
.status-note { color:#6edfd7; font-size:.68rem; margin-top:.18rem; }
.workflow-guide { display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:.55rem; margin:.15rem 0 .5rem; }
.workflow-step { display:flex; align-items:center; gap:.65rem; min-height:42px; padding:.48rem .7rem; border-radius:11px;
 border:1px solid rgba(57,242,160,.38); background:linear-gradient(135deg,rgba(8,45,52,.88),rgba(7,29,49,.9));
 box-shadow:inset 0 1px rgba(255,255,255,.035),0 0 18px rgba(57,242,160,.045); }
.workflow-number { display:grid; place-items:center; flex:0 0 25px; height:25px; border-radius:50%; color:#031a14;
 background:linear-gradient(135deg,#7dffc5,#27dfa0); font:800 .7rem 'IBM Plex Mono',monospace;
 box-shadow:0 0 12px rgba(57,242,160,.28); }
.workflow-text { color:#dffcf2; font-size:.72rem; font-weight:650; }
.kpi-grid { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:.8rem; margin:1rem 0; }
.kpi { padding:1rem 1.05rem; border-radius:15px; border:1px solid #285a78; background:linear-gradient(145deg,#0d2b46,#081e33);
 box-shadow:0 12px 30px rgba(0,0,0,.16),inset 0 1px rgba(255,255,255,.04); }
.kpi-label { color:#94aec2; font:600 .65rem 'IBM Plex Mono',monospace; letter-spacing:.09em; text-transform:uppercase; }
.kpi-value { color:#fff; font-size:1.34rem; font-weight:800; margin-top:.35rem; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.kpi-accent { height:2px; width:34px; margin-top:.65rem; background:linear-gradient(90deg,#4bdbd2,#168ec8); border-radius:3px; }
[data-testid="stFileUploader"] { background:#0b233a; border:1px dashed #37779d; border-radius:14px; padding:.5rem; }
.stTextInput input, [data-baseweb="select"] > div { border-radius:10px !important; }
.stTextInput input[aria-label="Usuario"],
.stTextInput input[aria-label="Usuario autorizado"] { text-transform:uppercase; letter-spacing:.045em; }
.login-shell { padding:.05rem .25rem .35rem; }
.login-kicker { color:#52d8cf; font:600 .68rem 'IBM Plex Mono',monospace; letter-spacing:.11em; text-transform:uppercase; }
.stColumn:has(.login-marker) [data-testid="stVerticalBlockBorderWrapper"] { min-height:410px; }
.snoopy-login-page { display:none; }
[data-testid="stAppViewContainer"]:has(.snoopy-login-page) {
 height:100vh; min-height:100vh; overflow:hidden;
 background:
  radial-gradient(circle at 50% 49%,rgba(17,103,143,.22),transparent 24%),
  radial-gradient(circle at 9% 30%,rgba(22,151,213,.14),transparent 28%),
  radial-gradient(circle at 92% 68%,rgba(72,215,208,.12),transparent 27%),
  linear-gradient(rgba(45,113,151,.07) 1px,transparent 1px),
  linear-gradient(90deg,rgba(45,113,151,.07) 1px,transparent 1px),
  linear-gradient(145deg,#020a14 0%,#06172a 48%,#03101d 100%) !important;
 background-size:auto,auto,auto,42px 42px,42px 42px,auto !important;
}
[data-testid="stAppViewContainer"]:has(.snoopy-login-page)::before,
[data-testid="stAppViewContainer"]:has(.snoopy-login-page)::after {
 content:""; position:fixed; top:11vh; bottom:8vh; width:min(28vw,430px); pointer-events:none; opacity:.62;
 background:
  linear-gradient(180deg,transparent 0 12%,rgba(43,142,190,.18) 12% 12.35%,transparent 12.35% 25%,rgba(43,142,190,.13) 25% 25.3%,transparent 25.3%),
  repeating-linear-gradient(90deg,transparent 0 24px,rgba(44,129,173,.11) 25px 26px),
  radial-gradient(ellipse at center,rgba(17,109,151,.15),transparent 66%);
 filter:drop-shadow(0 0 24px rgba(13,132,184,.08)); z-index:0;
}
[data-testid="stAppViewContainer"]:has(.snoopy-login-page)::before { left:0; clip-path:polygon(0 0,88% 8%,72% 100%,0 92%); }
[data-testid="stAppViewContainer"]:has(.snoopy-login-page)::after { right:0; transform:scaleX(-1); clip-path:polygon(0 0,88% 8%,72% 100%,0 92%); }
[data-testid="stAppViewContainer"]:has(.snoopy-login-page) .block-container {
 position:relative; z-index:1; width:100%; max-width:1280px; height:100vh; min-height:100vh;
 margin:0 auto !important; padding:clamp(18px,3.6vh,40px) 1.2rem 1rem !important;
}
[data-testid="stAppViewContainer"]:has(.snoopy-login-page) .block-container > div:first-child {
 min-height:calc(100vh - clamp(36px,7.2vh,80px) - 1rem); display:flex; flex-direction:column; justify-content:center;
}
.snoopy-login-brand { margin:0 auto clamp(18px,3vh,32px); text-align:center; }
.snoopy-login-name {
 color:#9fdce8; font:800 clamp(2.3rem,4.5vw,4.15rem) 'IBM Plex Mono',monospace; line-height:1;
 letter-spacing:.1em; text-shadow:0 3px 18px rgba(0,0,0,.8),0 0 24px rgba(46,184,207,.18);
}
.snoopy-login-name span { color:#52d8cf; }
.stColumn:has(.snoopy-login-card) [data-testid="stVerticalBlockBorderWrapper"] {
 min-height:0 !important; border:1px solid rgba(66,129,165,.82); border-radius:20px;
 background:linear-gradient(145deg,rgba(7,27,49,.94),rgba(3,15,29,.91));
 box-shadow:0 28px 72px rgba(0,0,0,.55),0 0 28px rgba(39,144,185,.10),inset 0 1px rgba(255,255,255,.055);
 backdrop-filter:blur(13px);
}
.stColumn:has(.snoopy-login-card) [data-testid="stVerticalBlockBorderWrapper"] > div { padding:1.25rem 1.55rem 1.35rem; }
.snoopy-login-card { display:none; }
.stColumn:has(.snoopy-login-card) .stTextInput { margin-bottom:-.15rem; }
.stColumn:has(.snoopy-login-card) .stTextInput label { color:#bdd0dc; font-weight:600; }
.stColumn:has(.snoopy-login-card) .stTextInput input {
 min-height:3rem; color:#eef9ff; background:rgba(4,18,35,.9); border:1px solid #2b526e; border-radius:10px !important;
 box-shadow:inset 0 1px rgba(255,255,255,.025);
}
.stColumn:has(.snoopy-login-card) .stTextInput input:focus {
 border-color:#52d8cf !important; box-shadow:0 0 0 2px rgba(82,216,207,.16),0 0 16px rgba(82,216,207,.08) !important;
}
.stColumn:has(.snoopy-login-card) .stButton>button,
.stColumn:has(.snoopy-login-card) .stFormSubmitButton>button {
 margin-top:.2rem; letter-spacing:.08em; text-transform:uppercase;
 background:linear-gradient(90deg,#117fa8 0%,#12aaa4 55%,#2ac8a5 100%) !important;
 color:#f4ffff !important; border:1px solid rgba(104,231,218,.72) !important;
 box-shadow:0 0 0 1px rgba(82,216,207,.10),0 0 20px rgba(40,192,183,.16) !important;
}
.stColumn:has(.snoopy-login-card) .stButton>button:hover,
.stColumn:has(.snoopy-login-card) .stFormSubmitButton>button:hover {
 filter:brightness(1.1); border-color:#92f4e9 !important;
 box-shadow:0 0 25px rgba(53,213,201,.28) !important;
}
.excel-visual { position:relative; min-height:410px; height:100%; border-radius:18px; border:1px solid #285b79; overflow:hidden;
 background:radial-gradient(circle at 50% 38%,rgba(31,190,185,.17),transparent 48%),linear-gradient(145deg,#0c2943,#071c30); padding:1rem; }
.excel-visual:after { content:""; position:absolute; inset:0; background:linear-gradient(90deg,transparent 52%,rgba(7,28,48,.95) 96%); pointer-events:none; }
.doc-tag { color:#66ddd5; font:600 .66rem 'IBM Plex Mono',monospace; letter-spacing:.1em; text-transform:uppercase; }
.sheet { position:absolute; left:9%; top:18%; width:82%; transform:perspective(700px) rotateY(8deg) rotate(-1.5deg);
 border:1px solid rgba(111,213,209,.45); border-radius:10px; background:rgba(9,35,57,.88); box-shadow:0 24px 45px rgba(0,0,0,.34); overflow:hidden; }
.sheet-head,.sheet-row { display:grid; grid-template-columns:.7fr 1.65fr .8fr .8fr; }
.sheet-head span { padding:.55rem .42rem; background:#0e806f; color:#eafffd; font:600 .54rem 'IBM Plex Mono',monospace; }
.sheet-row span { padding:.48rem .42rem; border-right:1px solid rgba(56,105,131,.5); border-top:1px solid rgba(56,105,131,.42);
 color:#aec6d7; font:500 .52rem 'IBM Plex Mono',monospace; white-space:nowrap; overflow:hidden; }
.flow-arrow { position:absolute; right:5%; bottom:9%; color:#65e3d9; font:600 .66rem 'IBM Plex Mono',monospace; letter-spacing:.08em; z-index:2; }
.side-disclaimer { position:relative; min-height:410px; height:100%; padding:1rem; border-radius:18px; border:1px solid #285b79; overflow:hidden;
 background:radial-gradient(circle at 50% 38%,rgba(31,190,185,.17),transparent 48%),linear-gradient(215deg,#0c2943,#071c30);
 color:#afc0cf; font-size:.72rem; line-height:1.52; box-shadow:inset 0 1px 0 rgba(255,255,255,.035); }
.side-disclaimer:after { content:""; position:absolute; inset:0; background:linear-gradient(-90deg,transparent 52%,rgba(7,28,48,.95) 96%); pointer-events:none; z-index:1; }
.side-disclaimer .legal-title { margin-bottom:.48rem; }
.side-disclaimer strong { color:#fff; }
.disclaimer-head { display:flex; align-items:flex-start; justify-content:space-between; gap:.65rem; margin-bottom:.45rem; }
.disclaimer-author { color:#e9c968; font:600 .62rem 'IBM Plex Mono',monospace; white-space:nowrap; text-align:right; }
.econ-visual { position:absolute; inset:0; padding:1rem; overflow:hidden; z-index:2; }
.econ-title { color:#66ddd5; font:600 .66rem 'IBM Plex Mono',monospace; letter-spacing:.1em; text-transform:uppercase; }
.ledger-mini { position:absolute; right:9%; top:18%; width:82%; border:1px solid rgba(111,213,209,.45); border-radius:10px; overflow:hidden;
 transform:perspective(700px) rotateY(-8deg) rotate(1.5deg); transform-origin:center; background:rgba(9,35,57,.88);
 box-shadow:0 24px 45px rgba(0,0,0,.34); }
.ledger-head,.ledger-row { display:grid; grid-template-columns:.72fr 1.55fr .72fr .72fr; }
.ledger-head span { padding:.55rem .42rem; background:#0e806f; color:#eafffd; font:600 .54rem 'IBM Plex Mono',monospace; }
.ledger-row span { padding:.48rem .42rem; border-top:1px solid rgba(56,105,131,.42); border-right:1px solid rgba(56,105,131,.5); color:#aec6d7; font:500 .52rem 'IBM Plex Mono',monospace; white-space:nowrap; overflow:hidden; }
.econ-equality { position:absolute; left:9%; right:9%; top:57%; display:flex; justify-content:space-between; align-items:center; color:#73e1d9;
 font:700 .55rem 'IBM Plex Mono',monospace; letter-spacing:.06em; }
.econ-concepts { position:absolute; left:9%; right:9%; top:65%; color:#8eacc0; font:600 .49rem 'IBM Plex Mono',monospace; letter-spacing:.08em; }
.econ-flow { position:absolute; left:9%; right:9%; bottom:9%; color:#f0c75b; font:700 .6rem 'IBM Plex Mono',monospace; letter-spacing:.08em; text-align:left; }
.world-banner { position:relative; margin-top:-.1rem; height:74px; border-radius:16px; overflow:hidden; border:1px solid #2b6380;
 background:linear-gradient(180deg,rgba(101,190,225,.17) 0 33%,rgba(238,248,252,.10) 33% 66%,rgba(101,190,225,.17) 66%);
 box-shadow:inset 0 1px rgba(255,255,255,.05),0 12px 30px rgba(0,0,0,.14); }
.world-banner:before,.world-banner:after { content:""; position:absolute; top:0; bottom:0; width:90px; z-index:2; pointer-events:none; }
.world-banner:before { left:0; background:linear-gradient(90deg,#082039,transparent); }
.world-banner:after { right:0; background:linear-gradient(-90deg,#082039,transparent); }
.world-track { position:absolute; inset:0; display:flex; align-items:center; justify-content:center; animation:champions 8s ease-in-out infinite alternate; }
.champions { display:flex; align-items:center; gap:1.8rem; white-space:nowrap; }
.champion-title { color:#dcecf5; font:600 .68rem 'IBM Plex Mono',monospace; letter-spacing:.14em; }
.star-block { display:flex; align-items:center; gap:.45rem; color:#f0c75b; }
.star { font-size:1.75rem; line-height:1; filter:drop-shadow(0 0 7px rgba(240,199,91,.35)); }
.star-year { color:#f3d77f; font:600 .67rem 'IBM Plex Mono',monospace; }
@keyframes champions { from { transform:translateX(-9%); } to { transform:translateX(9%); } }
.stButton>button, .stDownloadButton>button { border-radius:9px; font-weight:700; min-height:2.65rem; }
.stButton>button[kind="primary"], .stDownloadButton>button[kind="primary"] {
 background:linear-gradient(90deg,#087fba 0%,#0db58e 58%,#39e99f 100%) !important; color:#fff !important;
 border:1px solid rgba(104,255,194,.72) !important; box-shadow:0 0 0 1px rgba(57,242,160,.12),0 0 18px rgba(57,242,160,.18) !important; }
.stButton>button[kind="primary"]:hover, .stDownloadButton>button[kind="primary"]:hover {
 filter:brightness(1.12); border-color:#86ffca !important; box-shadow:0 0 24px rgba(57,242,160,.32) !important; }
button:focus, button:focus-visible, input:focus, [data-baseweb="select"]>div:focus-within,
[data-testid="stFileUploader"]:focus-within { outline:none !important; border-color:#39f2a0 !important;
 box-shadow:0 0 0 2px rgba(57,242,160,.22) !important; }
input[type="checkbox"], input[type="radio"] { accent-color:#39f2a0 !important; }
[data-baseweb="radio"]>div:first-child { border-color:#39f2a0 !important; }
[data-baseweb="radio"]>div:first-child:after { background-color:#39f2a0 !important; }
hr { border-color:#20425d; }
@media(max-width:850px){.hero{grid-template-columns:1fr}.brand-lockup{text-align:left;min-width:0;align-self:auto}.brand-author{justify-content:flex-start}.hero-meta{text-align:left}.status-grid,.kpi-grid{grid-template-columns:1fr 1fr}.workflow-guide{grid-template-columns:1fr}.excel-visual,.side-disclaimer{min-height:340px}.world-track{animation:none}.snoopy-login-name{font-size:2.25rem}.stColumn:has(.snoopy-login-card) [data-testid="stVerticalBlockBorderWrapper"]>div{padding:1rem 1.1rem}}
@media(max-width:560px){.status-grid,.kpi-grid{grid-template-columns:1fr}}
</style>
""", unsafe_allow_html=True)


def _hero_image_uri() -> str:
    """Carga el fondo institucional sin exponer una ruta local al navegador."""
    image_path = Path(__file__).with_name("snoopy_accountant_hero.webp")
    try:
        encoded = base64.b64encode(image_path.read_bytes()).decode("ascii")
        return f"data:image/webp;base64,{encoded}"
    except OSError:
        return ""


_HERO_IMAGE = _hero_image_uri()
if _HERO_IMAGE:
    st.markdown(f"""
    <style>
    .hero {{
      position:relative;
      isolation:isolate;
      background:
        linear-gradient(90deg,rgba(7,27,47,.99) 0%,rgba(7,27,47,.97) 43%,rgba(7,27,47,.55) 68%,rgba(7,27,47,.20) 100%),
        linear-gradient(180deg,rgba(7,27,47,.12),rgba(7,27,47,.58)),
        url('{_HERO_IMAGE}') right center/cover no-repeat;
    }}
    .hero:after {{
      content:""; position:absolute; inset:0; z-index:-1; pointer-events:none;
      background:radial-gradient(circle at 83% 34%,rgba(38,206,210,.16),transparent 30%);
    }}
    .hero > div:first-child {{ position:relative; z-index:2; }}
    .brand-lockup {{ position:relative; z-index:3; }}
    </style>
    """, unsafe_allow_html=True)


def _login_background_uri() -> str:
    """Carga el fondo de acceso institucional desde un recurso local de la aplicación."""
    image_path = Path(__file__).with_name("snoopy_login_background.webp")
    try:
        encoded = base64.b64encode(image_path.read_bytes()).decode("ascii")
        return f"data:image/webp;base64,{encoded}"
    except OSError:
        return ""


_LOGIN_BACKGROUND = _login_background_uri()
if _LOGIN_BACKGROUND:
    st.markdown(f"""
    <style>
    [data-testid="stAppViewContainer"]:has(.snoopy-login-page) {{
      background:
        linear-gradient(180deg,rgba(1,9,19,.18),rgba(1,9,19,.36)),
        radial-gradient(circle at 50% 48%,rgba(5,21,39,.10),rgba(2,11,23,.26) 44%,rgba(1,8,17,.44) 100%),
        url('{_LOGIN_BACKGROUND}') center center/cover no-repeat fixed !important;
    }}
    [data-testid="stAppViewContainer"]:has(.snoopy-login-page)::before,
    [data-testid="stAppViewContainer"]:has(.snoopy-login-page)::after {{ display:none !important; }}
    </style>
    """, unsafe_allow_html=True)


def now_ar() -> datetime:
    return datetime.now(TZ_AR)


def period_label() -> str:
    value = now_ar()
    return f"{MONTHS[value.month]} {value.year}"


def hero(subtitle: str, show_author: bool = True) -> None:
    author_html = (
        '<div class="brand-author"><span class="brand-xicon">X</span><span>@PamperoSur</span></div>'
        if show_author else ""
    )
    st.markdown(f"""
    <div class="hero">
      <div class="hero-copy"><div class="eyebrow">Intelligence workspace · Uso educativo</div>
      <h1>PDF bancario <span class="title-arrow">→</span> Excel normalizado</h1><p>{subtitle}</p></div>
      <div class="brand-lockup"><div class="brand-name">SNOOPY<span class="brand-ia">IA</span></div>
      {author_html}
      <div class="hero-meta">Corrientes · Argentina · {period_label()}</div></div>
    </div>""", unsafe_allow_html=True)


def money_ar(value: float) -> str:
    raw = f"{float(value):,.2f}"
    return "$ " + raw.replace(",", "X").replace(".", ",").replace("X", ".")


def datetime_ar(value) -> str:
    if not value or pd.isna(value):
        return "Sin ingresos"
    try:
        parsed = pd.to_datetime(value, utc=True).tz_convert(TZ_AR)
        return parsed.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(value)


def password_age_days(value) -> int | None:
    if not value or pd.isna(value):
        return None
    try:
        changed = pd.to_datetime(value, utc=True).tz_convert(TZ_AR)
        return max(0, (now_ar().date() - changed.date()).days)
    except Exception:
        return None


def clean_view(frame: pd.DataFrame) -> pd.DataFrame:
    view = frame.copy()
    if "Fecha" in view.columns:
        view["Fecha"] = pd.to_datetime(view["Fecha"], errors="coerce").dt.strftime("%d/%m/%Y")
    return view.where(pd.notna(view), "")


@st.cache_resource
def db() -> Client:
    return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])


def headers() -> dict[str, str]:
    try:
        return {str(k).lower(): str(v) for k, v in st.context.headers.items()}
    except Exception:
        return {}


def client_info() -> dict[str, str]:
    h = headers()
    candidates = []
    for key in ("cf-connecting-ip", "true-client-ip", "x-real-ip", "x-forwarded-for"):
        candidates.extend(part.strip() for part in h.get(key, "").split(",") if part.strip())
    try:
        context_ip = str(st.context.ip_address or "").strip()
        if context_ip:
            candidates.append(context_ip)
    except Exception:
        pass
    ip = "N/D"
    for candidate in candidates:
        try:
            if ipaddress.ip_address(candidate).is_global:
                ip = candidate
                break
        except ValueError:
            continue
    if ip == "N/D" and candidates:
        ip = candidates[0]
    ua = h.get("user-agent", "N/D")
    info = {"ip_address": ip, "city": h.get("x-vercel-ip-city", "N/D"),
            "region": h.get("x-vercel-ip-country-region", "N/D"),
            "country": h.get("cf-ipcountry", h.get("x-vercel-ip-country", "N/D")),
            "user_agent": ua}
    if info["city"] == "N/D" and ip not in {"", "N/D"}:
        geo = geolocate(ip)
        info.update({k: geo.get(k, info[k]) for k in ("city", "region", "country")})
    return info


def simple_device(user_agent: str) -> str:
    """Convierte el user-agent técnico en una descripción breve y útil."""
    ua = str(user_agent or "")
    if not ua or ua == "N/D":
        return "N/D"
    if "HeadlessChrome" in ua:
        match = re.search(r"HeadlessChrome/(\d+)", ua)
        return f"Control automático · Chrome {match.group(1) if match else ''}".strip()
    browser = "Navegador"
    version = ""
    patterns = (("Edg/", "Edge"), ("OPR/", "Opera"), ("Chrome/", "Chrome"),
                ("Firefox/", "Firefox"), ("Version/", "Safari"))
    for token, name in patterns:
        match = re.search(re.escape(token) + r"(\d+)", ua)
        if match:
            browser, version = name, match.group(1)
            break
    if "iPhone" in ua:
        system = "iPhone"
    elif "iPad" in ua:
        system = "iPad"
    elif "Android" in ua:
        system = "Android"
    elif "Macintosh" in ua or "Mac OS X" in ua:
        system = "macOS"
    elif "Windows" in ua:
        system = "Windows"
    elif "Linux" in ua:
        system = "Linux"
    else:
        system = "Dispositivo desconocido"
    return f"{browser}{' ' + version if version else ''} · {system}"


@st.cache_data(ttl=3600, show_spinner=False)
def geolocate(ip: str) -> dict[str, str]:
    try:
        address = ipaddress.ip_address(ip)
        if address.is_private or address.is_loopback:
            return {}
        req = urllib.request.Request(f"https://ipwho.is/{ip}", headers={"User-Agent": "ExtractorBancario/2.0"})
        with urllib.request.urlopen(req, timeout=2) as response:
            import json
            data = json.loads(response.read().decode("utf-8"))
        if not data.get("success", True):
            return {}
        return {"city": data.get("city") or "N/D", "region": data.get("region") or "N/D",
                "country": data.get("country") or "N/D"}
    except Exception:
        return {}


def log_access(username: str, success: bool, event_type: str = "LOGIN") -> None:
    payload = {"username": username.lower().strip() or "(vacío)", "success": success,
               "event_type": event_type, **client_info()}
    try:
        db().table("access_logs").insert(payload).execute()
    except Exception:
        pass


def acceptance_required(username: str) -> bool:
    """Exige aceptación inicial, cada 10 ingresos y después de cambiar la clave."""
    normalized = username.lower().strip()
    if not normalized:
        return True
    try:
        rows = (db().table("access_logs")
                .select("event_type,success,created_at")
                .eq("username", normalized)
                .order("created_at", desc=True)
                .limit(100).execute().data or [])
        accepted_at = None
        password_reset_at = None
        successful_logins = 0
        for row in rows:
            event = str(row.get("event_type") or "")
            created = str(row.get("created_at") or "")
            if event == "TERMS_ACCEPTED" and accepted_at is None:
                accepted_at = created
            elif event == "PASSWORD_RESET" and password_reset_at is None:
                password_reset_at = created
            elif event == "LOGIN" and bool(row.get("success")):
                if accepted_at is None or created > accepted_at:
                    successful_logins += 1
        return accepted_at is None or successful_logins >= 9 or (
            password_reset_at is not None and password_reset_at > accepted_at
        )
    except Exception:
        return True


def register_page_open() -> None:
    """Registra una sola apertura por sesión, incluso antes del inicio de sesión."""
    if not st.session_state.get("_page_open_logged"):
        log_access("(visitante)", True, "PAGE_OPEN")
        st.session_state._page_open_logged = True


def authenticate(username: str, password: str):
    result = db().rpc("verify_app_user", {"p_username": username, "p_password": password}).execute()
    rows = result.data or []
    if rows:
        log_access(username, True)
        user = rows[0]
        try:
            metadata = (db().table("app_users").select(
                "username,full_name,role,active,status,archived,password_changed_at,must_change_password"
            ).eq("username", username.lower().strip()).single().execute().data or {})
            user.update(metadata)
        except Exception:
            user.setdefault("status", "active")
            user.setdefault("archived", False)
            user.setdefault("must_change_password", False)
            user.setdefault("password_changed_at", None)
        return user
    log_access(username, False)
    return None


def gateway_user_authorized(username: str) -> bool:
    """Comprueba silenciosamente que el usuario exista y esté habilitado."""
    normalized = username.lower().strip()
    if not normalized:
        return False
    try:
        rows = (db().table("app_users")
                .select("username,active,status,archived")
                .eq("username", normalized)
                .limit(1).execute().data or [])
        if not rows:
            return False
        account = rows[0]
        status = str(account.get("status") or "active").lower().strip()
        blocked_states = {"blocked", "archived", "disabled", "paused", "inactive"}
        return bool(account.get("active", False)) and not bool(account.get("archived", False)) \
            and status not in blocked_states
    except Exception:
        return False


def academic_notice() -> None:
    st.markdown(f"""
    <div class="legal"><div class="legal-title">Descargo de responsabilidad</div>
    <b>Emprendedurismo (IA) - Corrientes · {APP_VERSION} · USO EXCLUSIVAMENTE EDUCATIVO</b><br><br>
    {DISCLAIMER}</div>""", unsafe_allow_html=True)


def excel_preview() -> None:
    st.markdown("""
    <div class="excel-visual"><div class="doc-tag">PDF RAW DATA → EXCEL</div>
      <div class="sheet">
        <div class="sheet-head"><span>FECHA</span><span>CONCEPTO</span><span>CRÉDITO</span><span>DÉBITO</span></div>
        <div class="sheet-row"><span>03/08</span><span>TRANSFERENCIA CUIT 20...</span><span>125.000</span><span>—</span></div>
        <div class="sheet-row"><span>05/08</span><span>DEPÓSITO EFECTIVO</span><span>82.500</span><span>—</span></div>
        <div class="sheet-row"><span>11/08</span><span>PAGO PROVEEDOR</span><span>—</span><span>43.200</span></div>
        <div class="sheet-row"><span>18/08</span><span>TRANSFERENCIA RECIBIDA</span><span>310.800</span><span>—</span></div>
        <div class="sheet-row"><span>22/08</span><span>IMPUESTO DÉB./CRÉD.</span><span>—</span><span>1.865</span></div>
      </div><div class="flow-arrow">NORMALIZADO ✓</div>
    </div>""", unsafe_allow_html=True)


def side_disclaimer() -> None:
    st.markdown("""
    <div class="side-disclaimer">
      <div class="econ-visual"><div class="econ-title">El ADN de las Ciencias Económicas</div>
        <div class="ledger-mini"><div class="ledger-head"><span>FECHA</span><span>CUENTA</span><span>DEBE</span><span>HABER</span></div>
        <div class="ledger-row"><span>18/08</span><span>BANCO</span><span>125.000</span><span>—</span></div>
        <div class="ledger-row"><span>18/08</span><span>TRANSFERENCIAS</span><span>—</span><span>125.000</span></div>
        <div class="ledger-row"><span>31/08</span><span>IMPUESTOS</span><span>8.750</span><span>—</span></div></div>
        <div class="econ-equality"><span>DEBE = HABER</span><span>CONCILIADO ✓</span></div>
        <div class="econ-concepts">CUIT · IVA · IIBB · AUDITORÍA · CONTROL</div>
        <div class="econ-flow">DATOS → CONTROL → CONOCIMIENTO</div>
      </div>
    </div>""", unsafe_allow_html=True)


def champions_banner() -> None:
    st.markdown("""
    <div class="world-banner"><div class="world-track"><div class="champions">
      <div class="champion-title">ARGENTINA · TRES VECES CAMPEÓN DEL MUNDO</div>
      <div class="star-block"><span class="star">★</span><span class="star-year">1978</span></div>
      <div class="star-block"><span class="star">★</span><span class="star-year">1986</span></div>
      <div class="star-block"><span class="star">★</span><span class="star-year">2022</span></div>
      <div class="champion-title">ORGULLO · HISTORIA · IDENTIDAD</div>
    </div></div></div>""", unsafe_allow_html=True)


def gateway_screen() -> None:
    st.markdown('<span class="snoopy-login-page"></span>', unsafe_allow_html=True)
    st.markdown("""
    <div class="snoopy-login-brand">
      <div class="snoopy-login-name">SNOOPY <span>IA</span></div>
    </div>
    """, unsafe_allow_html=True)
    left, center, right = st.columns([1.15, 1, 1.15], gap="large")
    with center:
        with st.container(border=True):
            st.markdown('<span class="snoopy-login-card"></span>', unsafe_allow_html=True)
            with st.form("snoopy_gateway_form", clear_on_submit=False):
                gateway_username = st.text_input("Usuario autorizado", key="gateway_username")
                submitted = st.form_submit_button("Continuar", type="primary", width="stretch")
            if submitted:
                normalized = gateway_username.lower().strip()
                authorized = gateway_user_authorized(normalized)
                log_access(normalized or "(vacío)", authorized, "GATEWAY_ACCESS")
                if authorized:
                    st.session_state["snoopy_gateway_passed"] = True
                    st.session_state["snoopy_gateway_username"] = normalized
                    st.rerun()
                else:
                    st.error("Acceso no habilitado.")


def login_screen() -> None:
    if not st.session_state.get("snoopy_gateway_passed", False):
        gateway_screen()
        return

    hero("Sistema inteligente de normalización bancaria", show_author=False)
    st.markdown(f"""<div class="status-grid">
      <div class="status-card"><div class="status-label">Cobertura</div><div class="status-value">8 entidades bancarias</div><div class="status-note">Lectores normalizados</div></div>
      <div class="status-card"><div class="status-label">Privacidad</div><div class="status-value">Procesamiento temporal</div><div class="status-note">Los PDF no se almacenan</div></div>
      <div class="status-card"><div class="status-label">Robustez</div><div class="status-value">PDF de gran volumen</div><div class="status-note">Procesamiento página por página</div></div>
      <div class="status-card"><div class="status-label">Fiabilidad</div><div class="status-value">Salida normalizada</div><div class="status-note">Excel, CSV y control de filas</div></div>
    </div>""", unsafe_allow_html=True)
    visual, center, legal = st.columns(3, gap="medium")
    with visual:
        excel_preview()
    with center:
        with st.container(border=True):
            st.markdown('<span class="login-marker"></span>', unsafe_allow_html=True)
            st.markdown('<div class="login-kicker">Acceso seguro · Control de usuarios</div>', unsafe_allow_html=True)
            st.subheader("Ingreso de usuarios")
            username = st.text_input("Usuario").strip().lower()
            password = st.text_input("Clave", type="password")
            must_accept = acceptance_required(username)
            accepted = True
            if must_accept:
                accepted = st.checkbox(
                    "He leído y acepto el uso exclusivamente educativo, el descargo de responsabilidad y el registro de acceso."
                )
                st.caption("Esta confirmación se solicita al primer ingreso, cada 10 accesos o después de cambiar la clave.")
            if st.button("Ingresar", type="primary", width="stretch", disabled=not accepted):
                try:
                    user = authenticate(username, password)
                    if user:
                        if must_accept:
                            log_access(username, True, "TERMS_ACCEPTED")
                        st.session_state.user = user
                        st.rerun()
                    else:
                        st.error("Usuario, clave o estado incorrecto.")
                except Exception as exc:
                    st.error(f"No fue posible conectar con la base de usuarios: {exc}")
    with legal:
        side_disclaimer()
    champions_banner()


def classify_origin(text: str) -> str:
    value = str(text).upper()
    rules = [("DEBIN", "DEBIN"), ("DEPOS", "Depósito"), ("EFECTIVO", "Efectivo"),
             ("CHEQUE", "Cheque"), ("POSNET", "Cobranza/POSNET"), ("TARJ", "Tarjeta"),
             ("TRANSF", "Transferencia"), ("COELS", "Transferencia"),
             ("TR INTER", "Transferencia"), ("CRED TR", "Transferencia")]
    return next((label for token, label in rules if token in value), "Otro/N.D.")


def analyze_movements(frame: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = frame.copy()
    df["Año"] = df["Fecha"].dt.year
    df["Mes número"] = df["Fecha"].dt.month
    df["Mes"] = df["Mes número"].map(lambda x: MONTHS[int(x)] if pd.notna(x) else "")
    df["Año-Mes"] = df["Fecha"].dt.strftime("%Y-%m")
    df["CUIT/DNI detectado"] = df["Concepto"].astype(str).str.extract(r"(?<!\d)(\d{11})(?!\d)", expand=False).fillna("N/D")
    df["Procedencia"] = df["Concepto"].map(classify_origin)
    df["Banco receptor"] = df["Banco"]
    df["Banco origen"] = "N/D"
    df["Nombre/Procedencia detectada"] = df.apply(extract_name, axis=1)
    if "Cuenta" not in df.columns:
        df["Cuenta"] = "N/D"
    df["Cuenta"] = df["Cuenta"].fillna("").astype(str).str.strip().replace({"": "N/D", "nan": "N/D", "None": "N/D"})
    preferred = ["Banco", "Cuenta", "Fecha", "Concepto", "Crédito", "Débito", "CUIT/DNI detectado",
                 "Nombre/Procedencia detectada", "Procedencia", "Banco receptor", "Banco origen",
                 "Operación", "Página", "Año", "Mes", "Año-Mes", "Origen", "Código trx", "Saldo"]
    df = df[[c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]]
    credits = df[pd.to_numeric(df["Crédito"], errors="coerce").fillna(0) > 0].copy()
    monthly = credits.groupby("Año-Mes", as_index=False).agg(
        Acreditaciones=("Crédito", "size"), Total_acreditado=("Crédito", "sum")
    )
    return df, credits, monthly


def split_movement_periods(full: pd.DataFrame, mode: str,
                           months_per_file: int = 6) -> list[tuple[str, pd.DataFrame]]:
    """Split normalized movements without deduplicating or cutting PDF pages."""
    if full.empty or mode == "Un Excel completo":
        return [("completo", full.copy())]
    work = full.copy()
    work["Fecha"] = pd.to_datetime(work["Fecha"], errors="coerce")
    work = work[work["Fecha"].notna()].copy()
    if work.empty:
        return [("sin_fecha", full.copy())]
    if mode == "Separar el Excel por año":
        return [
            (str(int(year)), rows.copy())
            for year, rows in work.groupby(work["Fecha"].dt.year, sort=True)
        ]

    months_per_file = max(1, min(int(months_per_file), 12))
    month_index = work["Fecha"].dt.year * 12 + work["Fecha"].dt.month - 1
    first_month = int(month_index.min())
    work["_Bloque salida"] = ((month_index - first_month) // months_per_file).astype(int)
    groups = []
    for _, rows in work.groupby("_Bloque salida", sort=True):
        rows = rows.drop(columns=["_Bloque salida"])
        start = rows["Fecha"].min().strftime("%Y-%m")
        end = rows["Fecha"].max().strftime("%Y-%m")
        groups.append((f"{start}_a_{end}", rows.copy()))
    return groups


def extract_name(row) -> str:
    text = re.sub(r"\s+", " ", str(row.get("Concepto", ""))).strip()
    identifier = str(row.get("CUIT/DNI detectado", "N/D"))
    if identifier != "N/D" and identifier in text:
        tail = text.split(identifier, 1)[1].strip(" -:/")
        return tail[:90] if tail else "N/D"
    return "N/D"


def _balance_control(full: pd.DataFrame) -> pd.DataFrame:
    """Reconcile opening balance, movements and closing balance for each account."""
    columns = ["Cuenta", "Saldo inicial", "Total créditos", "Total débitos",
               "Saldo final calculado", "Saldo final informado", "Diferencia",
               "Filas con diferencia", "Resultado"]
    if full.empty:
        return pd.DataFrame(columns=columns)

    work = full.copy()
    if "Cuenta" not in work.columns:
        work["Cuenta"] = "N/D"
    work["Cuenta"] = work["Cuenta"].fillna("").astype(str).str.strip().replace(
        {"": "N/D", "nan": "N/D", "None": "N/D"}
    )
    for name in ("Crédito", "Débito", "Saldo"):
        work[name] = pd.to_numeric(work.get(name), errors="coerce")

    results = []
    tolerance = 0.02
    for account, group in work.groupby("Cuenta", sort=False, dropna=False):
        group = group.reset_index(drop=True)
        credit = group["Crédito"].fillna(0.0)
        debit = group["Débito"].fillna(0.0)
        balances = group["Saldo"]
        informed_positions = [index for index, value in balances.items() if pd.notna(value)]
        total_credit = float(credit.sum())
        total_debit = float(debit.sum())

        if not informed_positions:
            results.append({"Cuenta": account, "Saldo inicial": None,
                            "Total créditos": total_credit, "Total débitos": total_debit,
                            "Saldo final calculado": None, "Saldo final informado": None,
                            "Diferencia": None, "Filas con diferencia": None,
                            "Resultado": "SIN SALDOS INFORMADOS"})
            continue

        first_position = informed_positions[0]
        last_position = informed_positions[-1]
        opening = float(balances.iloc[first_position]) - float(credit.iloc[:first_position + 1].sum()) + float(debit.iloc[:first_position + 1].sum())
        informed_closing = float(balances.iloc[last_position])
        calculated_closing = opening + total_credit - total_debit
        difference = calculated_closing - informed_closing

        row_differences = 0
        previous_position = first_position
        previous_balance = float(balances.iloc[first_position])
        for position in informed_positions[1:]:
            expected = previous_balance + float(credit.iloc[previous_position + 1:position + 1].sum()) - float(debit.iloc[previous_position + 1:position + 1].sum())
            if abs(expected - float(balances.iloc[position])) > tolerance:
                row_differences += 1
            previous_position = position
            previous_balance = float(balances.iloc[position])

        result = "OK" if abs(difference) <= tolerance and row_differences == 0 else "REVISAR"
        results.append({"Cuenta": account, "Saldo inicial": opening,
                        "Total créditos": total_credit, "Total débitos": total_debit,
                        "Saldo final calculado": calculated_closing,
                        "Saldo final informado": informed_closing, "Diferencia": difference,
                        "Filas con diferencia": row_differences, "Resultado": result})
    return pd.DataFrame(results, columns=columns)


def _account_sheet_name(account: str, used: set[str]) -> str:
    """Build a valid, unique Excel sheet name (maximum 31 characters)."""
    cleaned = re.sub(r"[\\/*?:\[\]]", "-", str(account)).strip() or "N-D"
    base = f"Cuenta {cleaned}"[:31]
    candidate = base
    suffix = 2
    while candidate in used:
        tail = f" ({suffix})"
        candidate = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def export_workbook(bank: str, full: pd.DataFrame, credits: pd.DataFrame,
                    grouped: pd.DataFrame, monthly: pd.DataFrame, rejected: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    balance_control = _balance_control(full)
    status = "Sin movimientos en el período" if full.empty else "Correcto"
    accounts = full.attrs.get("accounts", [])
    if not accounts and "Cuenta" in full.columns and not full.empty:
        accounts = full["Cuenta"].dropna().astype(str).drop_duplicates().tolist()
    control = pd.DataFrame({
        "Control": ["Aplicación", "Finalidad", "Autoría", "Entorno", "Emisión", "Banco",
                    "Estado", "Cuenta(s)", "Movimientos", "Acreditaciones",
                    "Total acreditado", "Filas a revisar"],
        "Resultado": [f"Extractor Bancario IA - {APP_VERSION}", "Uso exclusivamente educativo",
                      AUTHOR_CREDIT, "Emprendedurismo (IA) - Corrientes", period_label(), bank,
                      status, ", ".join(accounts) if accounts else "N/D", len(full), len(credits),
                      credits["Crédito"].sum(), len(rejected)],
    })
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        full.to_excel(writer, sheet_name="Movimientos", index=False)
        credits.to_excel(writer, sheet_name="Créditos", index=False)
        grouped.to_excel(writer, sheet_name="Agrupación créditos", index=False)
        monthly.to_excel(writer, sheet_name="Resumen mensual", index=False)
        control.to_excel(writer, sheet_name="Control", index=False)
        balance_control.to_excel(writer, sheet_name="Control de saldos", index=False)
        rejected.to_excel(writer, sheet_name="Revisar", index=False)
        if "Cuenta" in full.columns:
            account_values = full["Cuenta"].fillna("").astype(str).str.strip().replace(
                {"": "N/D", "nan": "N/D", "None": "N/D"}
            )
            if account_values.nunique(dropna=False) > 1:
                used_names = {sheet.title for sheet in writer.book.worksheets}
                for account in account_values.drop_duplicates():
                    account_rows = full.loc[account_values.eq(account)].copy()
                    account_rows.to_excel(
                        writer, sheet_name=_account_sheet_name(account, used_names), index=False
                    )
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="123A59")
                cell.font = Font(color="FFFFFF", bold=True)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or "")) for c in col) + 2, 55)
            headers = {cell.value: cell.column for cell in ws[1]}
            if "Fecha" in headers:
                for column in ws.iter_cols(min_col=headers["Fecha"], max_col=headers["Fecha"], min_row=2):
                    for cell in column:
                        cell.number_format = "dd/mm/yyyy"
            for name in ("Crédito", "Débito", "Saldo", "Total_acreditado", "Saldo inicial",
                         "Total créditos", "Total débitos", "Saldo final calculado",
                         "Saldo final informado", "Diferencia"):
                if name in headers:
                    for column in ws.iter_cols(min_col=headers[name], max_col=headers[name], min_row=2):
                        for cell in column:
                            cell.number_format = '#,##0.00;[Red]-#,##0.00'
    return output.getvalue()


def _safe_file_part(value: str) -> str:
    """Return a portable filename component for ZIP downloads."""
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value)).strip("_")
    return cleaned[:70] or "extracto"


def _batch_report_workbook(report: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        report.to_excel(writer, sheet_name="Informe del proceso", index=False)
        ws = writer.book["Informe del proceso"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="123A59")
            cell.font = Font(color="FFFFFF", bold=True)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(cell.value or "")) for cell in col) + 2, 65)
        headers = {cell.value: cell.column for cell in ws[1]}
        for name in ("Total créditos", "Total débitos"):
            if name in headers:
                for column in ws.iter_cols(min_col=headers[name], max_col=headers[name], min_row=2):
                    for cell in column:
                        cell.number_format = '#,##0.00;[Red]-#,##0.00'
    return output.getvalue()


def multiple_extractor() -> None:
    files = st.file_uploader(
        "Seleccionar varios extractos PDF",
        type=["pdf"], accept_multiple_files=True, key="batch_uploader",
        help="Máximo 10 archivos, 60 MB por PDF y 150 MB por lote.",
    )
    token = tuple((item.name, item.size) for item in files) if files else ()
    if token != st.session_state.get("batch_token", ()):
        st.session_state.pop("batch_result", None)
        st.session_state.batch_token = token

    if files:
        total_mb = sum(item.size for item in files) / (1024 * 1024)
        st.caption(f"{len(files)} archivo(s) seleccionados · {total_mb:.1f} MB en total")
    option_col1, option_col2 = st.columns(2)
    batch_output_mode = option_col1.selectbox(
        "Organización de cada resultado",
        ["Un Excel completo", "Separar el Excel por año", "Separar por bloques de meses"],
        key="batch_output_mode",
    )
    months_per_file = 6
    if batch_output_mode == "Separar por bloques de meses":
        months_per_file = option_col1.number_input(
            "Meses por archivo", min_value=1, max_value=12, value=6, step=1,
            key="batch_months_per_file",
        )
    consolidate = option_col2.checkbox(
        "Generar también un Excel consolidado", value=True,
        help="Une todos los movimientos sin eliminar operaciones que parezcan repetidas.",
    )
    if files and st.button("Convertir lote", type="primary", width="stretch"):
        if len(files) > 10:
            st.error("El lote admite como máximo 10 PDF.")
            return
        oversized = [item.name for item in files if item.size > 60 * 1024 * 1024]
        if oversized:
            st.error(f"Superan 60 MB: {', '.join(oversized)}")
            return
        if sum(item.size for item in files) > 150 * 1024 * 1024:
            st.error("El lote completo supera el límite operativo de 150 MB.")
            return

        report_rows = []
        archive_buffer = io.BytesIO()
        success_count = 0
        successful_frames: list[pd.DataFrame] = []
        consolidated_rejected: list[pd.DataFrame] = []
        progress_bar = st.progress(0.0, text="Preparando el lote…")
        with zipfile.ZipFile(archive_buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for file_index, item in enumerate(files):
                try:
                    def update_batch_progress(done: int, total: int, index: int = file_index) -> None:
                        fraction = (index + (done / max(total, 1))) / len(files)
                        progress_bar.progress(
                            min(fraction, 1.0),
                            text=f"Archivo {index + 1} de {len(files)} · página {done} de {total}",
                        )

                    bank, frame, rejected = parse_pdf(item.getvalue(), "Automático", update_batch_progress)
                    if frame.empty and bank == "Desconocido":
                        raise ValueError("No se pudo reconocer el banco ni detectar movimientos.")
                    full, credits, monthly = analyze_movements(frame)
                    grouped = credits.groupby(
                        ["CUIT/DNI detectado", "Nombre/Procedencia detectada", "Procedencia", "Banco receptor"],
                        dropna=False, as_index=False,
                    ).agg(Acreditaciones=("Crédito", "size"),
                          Total_acreditado=("Crédito", "sum")).sort_values("Total_acreditado", ascending=False)
                    base_name = (
                        f"{file_index + 1:02d}_{_safe_file_part(bank)}_"
                        f"{_safe_file_part(Path(item.name).stem)}"
                    )
                    generated_names = []
                    period_groups = split_movement_periods(
                        full, batch_output_mode, int(months_per_file)
                    )
                    if batch_output_mode != "Un Excel completo" and not full.empty:
                        for period_name, period_rows in period_groups:
                            year_full, year_credits, year_monthly = analyze_movements(period_rows.copy())
                            year_grouped = year_credits.groupby(
                                ["CUIT/DNI detectado", "Nombre/Procedencia detectada", "Procedencia", "Banco receptor"],
                                dropna=False, as_index=False,
                            ).agg(Acreditaciones=("Crédito", "size"),
                                  Total_acreditado=("Crédito", "sum")).sort_values(
                                      "Total_acreditado", ascending=False)
                            excel_name = f"{base_name}_{period_name}_normalizado.xlsx"
                            archive.writestr(
                                excel_name,
                                export_workbook(bank, year_full, year_credits, year_grouped,
                                                year_monthly, rejected),
                            )
                            generated_names.append(excel_name)
                    else:
                        excel_name = f"{base_name}_normalizado.xlsx"
                        archive.writestr(
                            excel_name,
                            export_workbook(bank, full, credits, grouped, monthly, rejected),
                        )
                        generated_names.append(excel_name)

                    consolidated_part = frame.copy()
                    consolidated_part["Archivo origen"] = item.name
                    consolidated_part.attrs = frame.attrs.copy()
                    successful_frames.append(consolidated_part)
                    if not rejected.empty:
                        review_part = rejected.copy()
                        review_part.insert(0, "Archivo origen", item.name)
                        consolidated_rejected.append(review_part)
                    success_count += 1
                    st.session_state.conversions_session += 1
                    log_access(st.session_state.user["username"], True, "PDF_PROCESSED")
                    report_rows.append({
                        "Archivo original": item.name, "Banco detectado": bank,
                        "Estado": "Sin movimientos" if full.empty else "Correcto",
                        "Movimientos": len(full), "Acreditaciones": len(credits),
                        "Total créditos": pd.to_numeric(full["Crédito"], errors="coerce").fillna(0).sum(),
                        "Total débitos": pd.to_numeric(full["Débito"], errors="coerce").fillna(0).sum(),
                        "Filas a revisar": len(rejected),
                        "Excel generado": "; ".join(generated_names), "Detalle": "",
                    })
                except MemoryError:
                    report_rows.append({
                        "Archivo original": item.name, "Banco detectado": "N/D", "Estado": "Error",
                        "Movimientos": 0, "Acreditaciones": 0, "Total créditos": 0,
                        "Total débitos": 0, "Filas a revisar": 0, "Excel generado": "",
                        "Detalle": "Memoria insuficiente; dividir el archivo por períodos.",
                    })
                except Exception as exc:
                    report_rows.append({
                        "Archivo original": item.name, "Banco detectado": "N/D", "Estado": "Error",
                        "Movimientos": 0, "Acreditaciones": 0, "Total créditos": 0,
                        "Total débitos": 0, "Filas a revisar": 0, "Excel generado": "",
                        "Detalle": str(exc)[:300],
                    })
                finally:
                    progress_bar.progress(
                        (file_index + 1) / len(files),
                        text=f"Procesados {file_index + 1} de {len(files)} archivos",
                    )
            if consolidate and successful_frames:
                non_empty_frames = [part for part in successful_frames if not part.empty]
                if non_empty_frames:
                    combined_frame = pd.concat(non_empty_frames, ignore_index=True, sort=False)
                    combined_full, combined_credits, combined_monthly = analyze_movements(combined_frame)
                    combined_grouped = combined_credits.groupby(
                        ["CUIT/DNI detectado", "Nombre/Procedencia detectada", "Procedencia", "Banco receptor"],
                        dropna=False, as_index=False,
                    ).agg(Acreditaciones=("Crédito", "size"),
                          Total_acreditado=("Crédito", "sum")).sort_values(
                              "Total_acreditado", ascending=False)
                    combined_review = (
                        pd.concat(consolidated_rejected, ignore_index=True, sort=False)
                        if consolidated_rejected else pd.DataFrame()
                    )
                    archive.writestr(
                        "00_CONSOLIDADO_extractos_normalizados.xlsx",
                        export_workbook("Consolidado", combined_full, combined_credits,
                                        combined_grouped, combined_monthly, combined_review),
                    )
            report = pd.DataFrame(report_rows)
            archive.writestr("informe_del_proceso.xlsx", _batch_report_workbook(report))
        st.session_state.batch_result = (archive_buffer.getvalue(), pd.DataFrame(report_rows), success_count)

    if "batch_result" in st.session_state:
        archive_bytes, report, success_count = st.session_state.batch_result
        failures = len(report) - success_count
        if success_count:
            st.success(f"Lote terminado: {success_count} Excel generados y {failures} archivo(s) con error.")
        else:
            st.error("No se pudo generar ningún Excel. Consultá el informe del proceso.")
        st.dataframe(report, hide_index=True, width="stretch")
        st.download_button(
            "Bajar ZIP con los Excel",
            archive_bytes,
            "snoopy_3_0_extractos_normalizados.zip",
            "application/zip",
            type="primary", width="stretch",
        )
    elif not files:
        st.info("Seleccioná hasta 10 PDF. Cada extracto generará su propio Excel normalizado.")


def excel_consolidator() -> None:
    """Join previously normalized Snoopy workbooks without deleting similar rows."""
    files = st.file_uploader(
        "Seleccionar Excel normalizados de Snoopy",
        type=["xlsx"], accept_multiple_files=True, key="excel_consolidator_uploader",
        help="Subí entre 2 y 20 archivos. Se utiliza la hoja Movimientos de cada uno.",
    )
    token = tuple((item.name, item.size) for item in files) if files else ()
    if token != st.session_state.get("excel_consolidator_token", ()):
        st.session_state.pop("excel_consolidated_result", None)
        st.session_state.excel_consolidator_token = token

    if files:
        st.caption(
            f"{len(files)} archivo(s) seleccionados · "
            f"{sum(item.size for item in files) / (1024 * 1024):.1f} MB en total"
        )
    if files and st.button("Consolidar Excel", type="primary", width="stretch"):
        try:
            if not 2 <= len(files) <= 20:
                raise ValueError("Seleccioná entre 2 y 20 archivos Excel.")
            frames = []
            required = {"Fecha", "Concepto", "Crédito", "Débito"}
            for item in files:
                try:
                    frame = pd.read_excel(io.BytesIO(item.getvalue()), sheet_name="Movimientos")
                except ValueError as exc:
                    raise ValueError(f"{item.name}: no contiene la hoja Movimientos.") from exc
                missing = sorted(required - set(frame.columns))
                if missing:
                    raise ValueError(f"{item.name}: faltan columnas {', '.join(missing)}.")
                parsed_dates = pd.to_datetime(frame["Fecha"], errors="coerce", dayfirst=True)
                invalid_dates = int(parsed_dates.isna().sum())
                if invalid_dates:
                    raise ValueError(f"{item.name}: contiene {invalid_dates} fecha(s) inválida(s).")
                frame["Fecha"] = parsed_dates
                for column in ("Crédito", "Débito", "Saldo"):
                    if column not in frame.columns:
                        frame[column] = pd.NA
                    frame[column] = pd.to_numeric(frame[column], errors="coerce")
                if "Banco" not in frame.columns:
                    frame["Banco"] = "N/D"
                if "Cuenta" not in frame.columns:
                    frame["Cuenta"] = "N/D"
                frame["Archivo origen"] = item.name
                frames.append(frame)

            combined = pd.concat(frames, ignore_index=True, sort=False)
            # Deliberately do not call drop_duplicates: bank statements may have
            # legitimate equal-looking rows distinguished by operation/reference.
            full, credits, monthly = analyze_movements(combined)
            grouped = credits.groupby(
                ["CUIT/DNI detectado", "Nombre/Procedencia detectada", "Procedencia", "Banco receptor"],
                dropna=False, as_index=False,
            ).agg(Acreditaciones=("Crédito", "size"),
                  Total_acreditado=("Crédito", "sum")).sort_values(
                      "Total_acreditado", ascending=False)
            banks = sorted(
                value for value in full["Banco"].dropna().astype(str).unique()
                if value and value not in {"N/D", "nan"}
            )
            bank_label = banks[0] if len(banks) == 1 else "Consolidado"
            workbook = export_workbook(
                bank_label, full, credits, grouped, monthly, pd.DataFrame()
            )
            st.session_state.excel_consolidated_result = (
                workbook, len(full), len(credits),
                float(pd.to_numeric(full["Crédito"], errors="coerce").fillna(0).sum()),
                float(pd.to_numeric(full["Débito"], errors="coerce").fillna(0).sum()),
            )
        except Exception as exc:
            st.error(f"No se pudieron consolidar los Excel: {exc}")

    if "excel_consolidated_result" in st.session_state:
        workbook, movements, credit_count, total_credit, total_debit = (
            st.session_state.excel_consolidated_result
        )
        st.success(
            f"Consolidación terminada: {movements:,} movimientos y "
            f"{credit_count:,} acreditaciones.".replace(",", ".")
        )
        summary = pd.DataFrame({
            "Control": ["Archivos consolidados", "Movimientos", "Acreditaciones",
                        "Total créditos", "Total débitos", "Filas eliminadas"],
            "Resultado": [len(files), movements, credit_count, total_credit, total_debit, 0],
        })
        st.dataframe(summary, hide_index=True, width="stretch")
        st.download_button(
            "Bajar Excel consolidado", workbook,
            "snoopy_consolidado_extractos_normalizados.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", width="stretch",
        )
    elif not files:
        st.info("Subí los Excel normalizados que quieras unir en un único archivo.")


def extractor_page() -> None:
    hero("Sistema inteligente de normalización bancaria")
    if "conversions_session" not in st.session_state:
        st.session_state.conversions_session = 0
    st.markdown(f"""<div class="status-grid">
      <div class="status-card"><div class="status-label">Sesión actual</div><div class="status-value">{st.session_state.conversions_session} PDF procesados</div><div class="status-note">Sin persistencia documental</div></div>
      <div class="status-card"><div class="status-label">Seguridad</div><div class="status-value">Usuario autenticado</div><div class="status-note">Acceso y actividad registrados</div></div>
      <div class="status-card"><div class="status-label">Período operativo</div><div class="status-value">{period_label()}</div><div class="status-note">Actualización mensual automática</div></div>
    </div>""", unsafe_allow_html=True)
    st.markdown("""<div class="workflow-guide">
      <div class="workflow-step"><span class="workflow-number">1</span><span class="workflow-text">Elegí el banco o usá detección automática</span></div>
      <div class="workflow-step"><span class="workflow-number">2</span><span class="workflow-text">Subí uno o varios extractos bancarios en PDF</span></div>
      <div class="workflow-step"><span class="workflow-number">3</span><span class="workflow-text">Convertí, controlá y bajá el Excel</span></div>
    </div>""", unsafe_allow_html=True)
    mode = st.radio(
        "Modalidad", ["Un PDF", "Varios PDF", "Consolidar Excel"],
        horizontal=True, label_visibility="collapsed",
    )
    if mode == "Consolidar Excel":
        excel_consolidator()
        academic_notice()
        return
    if mode == "Varios PDF":
        multiple_extractor()
        academic_notice()
        return
    bank_choice = st.selectbox("Banco / lector", ["Automático", "BTF", "Patagonia", "BBVA", "Comafi", "Macro", "Galicia", "HSBC", "Santander"])
    output_mode = st.selectbox(
        "Organización del resultado",
        ["Un Excel completo", "Separar el Excel por año", "Separar por bloques de meses"],
        help="La separación anual se realiza después de leer todo el PDF, sin cortar operaciones entre páginas.",
    )
    months_per_file = 6
    if output_mode == "Separar por bloques de meses":
        months_per_file = st.number_input(
            "Meses por archivo", min_value=1, max_value=12, value=6, step=1,
            help="Ejemplo: 3 genera archivos trimestrales consecutivos; 6, semestrales.",
        )
    uploaded = st.file_uploader("Seleccionar un extracto PDF", type=["pdf"], accept_multiple_files=False)
    upload_token = (uploaded.name, uploaded.size) if uploaded else None
    previous_token = st.session_state.get("active_upload")
    if previous_token and upload_token != previous_token:
        st.session_state.pop("result", None)
        st.session_state.pop("downloads", None)
        st.session_state.pop("active_upload", None)
        if uploaded is None:
            st.toast("PDF retirado: datos y descargas eliminados de la sesión.", icon="🧹")
    if uploaded and st.button("Convertir extracto", type="primary", width="stretch"):
        with st.spinner("Extrayendo y normalizando movimientos página por página…"):
            try:
                pdf_bytes = uploaded.getvalue()
                if len(pdf_bytes) > 60 * 1024 * 1024:
                    raise ValueError("El archivo supera el límite operativo de 60 MB.")
                progress_bar = st.progress(0, text="Preparando el PDF…")
                def update_progress(done: int, total: int) -> None:
                    progress_bar.progress(done / total, text=f"Procesando página {done} de {total}")
                bank, frame, rejected = parse_pdf(pdf_bytes, bank_choice, update_progress)
                progress_bar.progress(1.0, text="Extracción terminada")
                if frame.empty and bank == "Desconocido":
                    raise ValueError("No se pudo reconocer el banco ni detectar movimientos.")
                full, credits, monthly = analyze_movements(frame)
                grouped = credits.groupby(
                    ["CUIT/DNI detectado", "Nombre/Procedencia detectada", "Procedencia", "Banco receptor"],
                    dropna=False, as_index=False
                ).agg(Acreditaciones=("Crédito", "size"), Total_acreditado=("Crédito", "sum")).sort_values("Total_acreditado", ascending=False)
                st.session_state.result = (
                    uploaded.name, bank, full, credits, grouped, monthly, rejected,
                    output_mode, int(months_per_file)
                )
                st.session_state.active_upload = upload_token
                st.session_state.conversions_session += 1
                st.session_state.pop("downloads", None)
                log_access(st.session_state.user["username"], True, "PDF_PROCESSED")
            except MemoryError:
                st.error("El servidor agotó memoria. El proceso se detuvo de forma controlada; probá dividir el PDF por períodos.")
            except Exception as exc:
                st.error(f"No se pudo convertir el PDF: {exc}")
    if "result" not in st.session_state:
        st.info("Subí un PDF. El documento y sus movimientos no se guardan en la base de usuarios.")
        academic_notice()
        return
    (_, bank, full, credits, grouped, monthly, rejected,
     output_mode, months_per_file) = st.session_state.result
    if full.empty:
        st.info(
            "El extracto fue reconocido correctamente y no contiene movimientos en el período. "
            "Podés descargar el Excel normalizado con su constancia en la pestaña Control."
        )
    total_credits = credits["Crédito"].sum()
    movement_count = f"{len(full):,}".replace(",", ".")
    credit_count = f"{len(credits):,}".replace(",", ".")
    st.markdown(f"""<div class="kpi-grid">
      <div class="kpi"><div class="kpi-label">Entidad detectada</div><div class="kpi-value">{bank}</div><div class="kpi-accent"></div></div>
      <div class="kpi"><div class="kpi-label">Movimientos</div><div class="kpi-value">{movement_count}</div><div class="kpi-accent"></div></div>
      <div class="kpi"><div class="kpi-label">Acreditaciones</div><div class="kpi-value">{credit_count}</div><div class="kpi-accent"></div></div>
      <div class="kpi"><div class="kpi-label">Total acreditado</div><div class="kpi-value" title="{money_ar(total_credits)}">{money_ar(total_credits)}</div><div class="kpi-accent"></div></div>
    </div>""", unsafe_allow_html=True)
    tabs = st.tabs(["Movimientos", "Análisis de créditos", "Agrupaciones", "Control", "Descargas"])
    with tabs[0]:
        preview = full.head(2000)
        if len(full) > len(preview):
            st.info(f"Vista optimizada: se muestran 2.000 de {len(full):,} movimientos. El Excel contiene el total.")
        st.dataframe(clean_view(preview), hide_index=True, width="stretch")
    with tabs[1]:
        col1, col2, col3 = st.columns(3)
        min_amount = col1.number_input("Crédito mínimo", min_value=0.0, value=0.0, step=10000.0)
        search = col2.text_input("Buscar CUIT, nombre o concepto")
        origins = col3.multiselect("Procedencia", sorted(credits["Procedencia"].dropna().unique()))
        view = credits[credits["Crédito"] >= min_amount]
        if search:
            mask = view.astype(str).apply(lambda col: col.str.contains(search, case=False, na=False)).any(axis=1)
            view = view[mask]
        if origins:
            view = view[view["Procedencia"].isin(origins)]
        ordered = view.sort_values("Crédito", ascending=False)
        if len(ordered) > 2000:
            st.info(f"Se muestran las primeras 2.000 de {len(ordered):,} acreditaciones filtradas.")
        st.dataframe(clean_view(ordered.head(2000)), hide_index=True, width="stretch")
    with tabs[2]:
        st.markdown("#### Por CUIT, persona, procedencia y banco receptor")
        st.dataframe(clean_view(grouped), hide_index=True, width="stretch")
        st.markdown("#### Por mes")
        st.dataframe(clean_view(monthly), hide_index=True, width="stretch")
    with tabs[3]:
        if rejected.empty:
            st.success("No se detectaron líneas dudosas.")
        else:
            st.warning(f"Hay {len(rejected)} líneas que requieren control.")
            st.dataframe(rejected, hide_index=True, width="stretch")
    with tabs[4]:
        st.caption("El archivo se prepara únicamente cuando lo solicitás, para reducir el uso de memoria del servidor.")
        if st.button("Preparar archivos de descarga", type="primary", width="stretch"):
            with st.spinner("Generando Excel y CSV…"):
                if output_mode != "Un Excel completo" and not full.empty:
                    yearly_buffer = io.BytesIO()
                    with zipfile.ZipFile(yearly_buffer, "w", compression=zipfile.ZIP_DEFLATED) as yearly_zip:
                        for period_name, period_rows in split_movement_periods(
                                full, output_mode, int(months_per_file)):
                            year_full, year_credits, year_monthly = analyze_movements(period_rows.copy())
                            year_grouped = year_credits.groupby(
                                ["CUIT/DNI detectado", "Nombre/Procedencia detectada", "Procedencia", "Banco receptor"],
                                dropna=False, as_index=False,
                            ).agg(Acreditaciones=("Crédito", "size"),
                                  Total_acreditado=("Crédito", "sum")).sort_values(
                                      "Total_acreditado", ascending=False)
                            yearly_zip.writestr(
                                f"{bank.lower()}_{period_name}_normalizado.xlsx",
                                export_workbook(bank, year_full, year_credits, year_grouped,
                                                year_monthly, rejected),
                            )
                    download_file = yearly_buffer.getvalue()
                    suffix = "por_anio" if output_mode == "Separar el Excel por año" else "por_periodos"
                    download_name = f"{bank.lower()}_normalizado_{suffix}.zip"
                    download_mime = "application/zip"
                else:
                    download_file = export_workbook(bank, full, credits, grouped, monthly, rejected)
                    download_name = f"{bank.lower()}_normalizado.xlsx"
                    download_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                st.session_state.downloads = (
                    download_file, download_name, download_mime,
                    credits.to_csv(index=False, sep=";", decimal=",", date_format="%d/%m/%Y").encode("utf-8-sig"),
                )
        if "downloads" in st.session_state:
            book, download_name, download_mime, csv = st.session_state.downloads
            d1, d2 = st.columns(2)
            d1.download_button("Bajar resultado", book, download_name, download_mime,
                               type="primary", width="stretch")
            d2.download_button("Descargar créditos CSV", csv, f"{bank.lower()}_creditos.csv", "text/csv", width="stretch")
    academic_notice()


def admin_users_page() -> None:
    st.title("Administración de usuarios")
    st.caption("Disponible exclusivamente para Administradores.")
    with st.expander("Crear usuario", expanded=True):
        with st.form("create_user", clear_on_submit=True):
            c1, c2 = st.columns(2)
            username = c1.text_input("Usuario").strip().lower()
            full_name = c2.text_input("Nombre completo").strip()
            password = c1.text_input("Clave inicial", type="password")
            role = c2.selectbox("Rol", ["Analista", "Administrador"])
            if st.form_submit_button("Crear usuario", type="primary", width="stretch"):
                try:
                    db().rpc("create_app_user", {"p_username": username, "p_full_name": full_name,
                                                  "p_password": password, "p_role": role,
                                                  "p_created_by": st.session_state.user["username"]}).execute()
                    st.success(f"Usuario {username} creado.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"No se pudo crear: {exc}")
    users = db().table("app_users").select(
        "username,full_name,role,active,status,archived,created_at,created_by,last_login,access_count,password_changed_at,must_change_password"
    ).order("username").execute().data or []
    user_df = pd.DataFrame(users).rename(columns={
        "username": "Usuario", "full_name": "Nombre completo", "role": "Rol", "active": "Activo",
        "created_at": "Creado", "created_by": "Creado por", "last_login": "Último ingreso",
        "access_count": "Cantidad de accesos", "status": "Estado", "archived": "Archivado",
        "password_changed_at": "Último cambio de clave", "must_change_password": "Cambio obligatorio"})
    if not user_df.empty:
        user_df["Creado"] = user_df["Creado"].map(datetime_ar)
        user_df["Último ingreso"] = user_df["Último ingreso"].map(datetime_ar)
        user_df["Antigüedad clave"] = user_df["Último cambio de clave"].map(
            lambda value: f"{password_age_days(value)} días" if password_age_days(value) is not None else "N/D"
        )
        user_df["Último cambio de clave"] = user_df["Último cambio de clave"].map(datetime_ar)
    st.dataframe(user_df, hide_index=True, width="stretch")
    editable = [u["username"] for u in users if u["username"] != "adm" and not u.get("archived", False)]
    if editable:
        st.markdown("#### Modificar usuario")
        target = st.selectbox("Usuario", editable)
        c1, c2, c3 = st.columns(3)
        new_role = c1.selectbox("Nuevo rol", ["Analista", "Administrador"], key="new_role")
        if c1.button("Cambiar rol", width="stretch"):
            db().table("app_users").update({"role": new_role}).eq("username", target).execute(); st.rerun()
        new_password = c2.text_input("Nueva clave", type="password")
        if c2.button("Restablecer clave", width="stretch", disabled=not new_password):
            db().rpc("reset_app_password", {"p_username": target, "p_new_password": new_password}).execute()
            db().table("app_users").update({
                "password_changed_at": now_ar().isoformat(), "must_change_password": True
            }).eq("username", target).execute()
            log_access(target, True, "PASSWORD_RESET")
            st.success("Clave temporal asignada. En el próximo ingreso deberá aceptar el descargo y crear una clave personal.")
        selected = next(u for u in users if u["username"] == target)
        if c3.button("Pausar" if selected["active"] else "Reactivar", width="stretch"):
            new_active = not selected["active"]
            db().table("app_users").update({
                "active": new_active, "status": "active" if new_active else "paused"
            }).eq("username", target).execute(); st.rerun()
        if c3.button("Archivar usuario", width="stretch"):
            db().table("app_users").update({
                "active": False, "status": "archived", "archived": True,
                "archived_at": now_ar().isoformat()
            }).eq("username", target).execute()
            log_access(target, True, "USER_ARCHIVED")
            st.rerun()


def password_page(forced: bool = False) -> None:
    st.title("Crear nueva clave" if forced else "Cambiar mi clave")
    if forced:
        st.warning("La clave utilizada es temporal. Para continuar debés crear una clave personal.")
    else:
        st.caption("La nueva clave reemplazará inmediatamente la clave actual.")
    with st.form("change_own_password", clear_on_submit=True):
        new_password = st.text_input("Nueva clave", type="password")
        confirmation = st.text_input("Repetir nueva clave", type="password")
        submitted = st.form_submit_button("Guardar nueva clave", type="primary", width="stretch")
        if submitted:
            if len(new_password) < 8:
                st.error("La nueva clave debe tener al menos 8 caracteres.")
            elif new_password != confirmation:
                st.error("Las claves ingresadas no coinciden.")
            else:
                username = st.session_state.user["username"]
                try:
                    db().rpc("reset_app_password", {
                        "p_username": username, "p_new_password": new_password
                    }).execute()
                    db().table("app_users").update({
                        "password_changed_at": now_ar().isoformat(), "must_change_password": False
                    }).eq("username", username).execute()
                    log_access(username, True, "PASSWORD_CHANGED")
                    st.session_state.user["must_change_password"] = False
                    st.session_state.user["password_changed_at"] = now_ar().isoformat()
                    st.success("Clave personal guardada correctamente.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"No se pudo actualizar la clave: {exc}")


def all_event_rows(event_type: str, batch_size: int = 1000) -> list[dict]:
    """Lee todos los eventos por páginas para no truncar los totales de gestión."""
    rows: list[dict] = []
    start = 0
    while True:
        batch = (db().table("access_logs").select("username,event_type")
                 .eq("event_type", event_type)
                 .range(start, start + batch_size - 1).execute().data or [])
        rows.extend(batch)
        if len(batch) < batch_size:
            break
        start += batch_size
    return rows


def management_page() -> None:
    st.title("Datos de gestión")
    st.caption("Indicadores acumulados por usuario. Panel exclusivo del administrador.")
    users = db().table("app_users").select(
        "username,full_name,role,active,status,archived,access_count,last_login,password_changed_at"
    ).order("username").execute().data or []
    pdf_rows = all_event_rows("PDF_PROCESSED")
    pdf_counts = pd.Series([str(row.get("username") or "").lower() for row in pdf_rows]).value_counts()
    total_pdfs = int(len(pdf_rows))
    records = []
    for item in users:
        username = str(item.get("username") or "").lower()
        accesses = int(item.get("access_count") or 0)
        pdfs = int(pdf_counts.get(username, 0))
        records.append({
            "Usuario": username.upper(),
            "Nombre": item.get("full_name") or "",
            "Rol": item.get("role") or "",
            "Estado": "Archivado" if item.get("archived") else ("Activo" if item.get("active") else "Pausado"),
            "Ingresos totales": accesses,
            "PDF procesados": pdfs,
            "PDF por ingreso": round(pdfs / accesses, 2) if accesses else 0.0,
            "% de PDF totales": round((pdfs / total_pdfs) * 100, 1) if total_pdfs else 0.0,
            "Último ingreso": datetime_ar(item.get("last_login")),
            "Antigüedad clave": password_age_days(item.get("password_changed_at")),
        })
    frame = pd.DataFrame(records)
    active_users = int((frame["Estado"] == "Activo").sum()) if not frame.empty else 0
    total_accesses = int(frame["Ingresos totales"].sum()) if not frame.empty else 0
    ratio = total_pdfs / total_accesses if total_accesses else 0
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Usuarios activos", active_users)
    k2.metric("Ingresos acumulados", total_accesses)
    k3.metric("PDF procesados", total_pdfs)
    k4.metric("PDF por ingreso", f"{ratio:.2f}")
    if frame.empty:
        st.info("Todavía no hay usuarios para mostrar.")
        return
    st.dataframe(frame, hide_index=True, width="stretch", column_config={
        "% de PDF totales": st.column_config.ProgressColumn(
            "% de PDF totales", min_value=0.0, max_value=100.0, format="%.1f%%"
        ),
        "Antigüedad clave": st.column_config.NumberColumn("Días desde cambio", format="%d días")
    })
    st.caption("PDF por ingreso = PDF procesados / ingresos totales. El porcentaje representa la participación de cada usuario sobre todos los PDF registrados.")


def audit_page() -> None:
    st.title("Auditoría de actividad")
    st.caption("Panel exclusivo del administrador. Los PDF y sus movimientos no se almacenan.")
    logs = db().table("access_logs").select(
        "username,event_type,success,ip_address,city,region,country,user_agent,created_at"
    ).order("created_at", desc=True).limit(1000).execute().data or []
    frame = pd.DataFrame(logs)
    if frame.empty:
        st.info("Todavía no hay actividad registrada.")
        return

    events = frame["event_type"].fillna("N/D").astype(str)
    successes = frame["success"].fillna(False).astype(bool)
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Eventos", f"{len(frame):,}".replace(",", "."))
    k2.metric("Ingresos correctos", int(((events == "LOGIN") & successes).sum()))
    k3.metric("Intentos fallidos", int(((events == "LOGIN") & ~successes).sum()))
    k4.metric("PDF procesados", int((events == "PDF_PROCESSED").sum()))
    valid_ips = frame["ip_address"].replace(["", "N/D", None], pd.NA).dropna()
    k5.metric("IP diferentes", valid_ips.nunique())

    users = sorted(frame["username"].dropna().astype(str).unique().tolist())
    event_options = sorted(events.unique().tolist())
    f1, f2, f3 = st.columns([1, 1.4, 1])
    selected_user = f1.selectbox("Usuario", ["Todos"] + users, key="audit_user")
    selected_events = f2.multiselect("Eventos", event_options, default=event_options, key="audit_events")
    selected_result = f3.selectbox("Resultado", ["Todos", "Correctos", "Fallidos"], key="audit_result")

    filtered = frame.copy()
    if selected_user != "Todos":
        filtered = filtered[filtered["username"].astype(str) == selected_user]
    filtered = filtered[filtered["event_type"].fillna("N/D").astype(str).isin(selected_events)]
    if selected_result != "Todos":
        wanted = selected_result == "Correctos"
        filtered = filtered[filtered["success"].fillna(False).astype(bool) == wanted]

    view = filtered.rename(columns={"username": "Usuario", "event_type": "Evento", "success": "Correcto",
                                    "ip_address": "IP", "user_agent": "Dispositivo",
                                    "created_at": "Fecha y hora"})
    view["Usuario"] = view["Usuario"].astype(str).str.upper()
    view["Fecha y hora"] = view["Fecha y hora"].map(datetime_ar)
    location_parts = filtered[["city", "region", "country"]].fillna("N/D").astype(str)
    view["Ubicación"] = location_parts.apply(
        lambda row: " · ".join(dict.fromkeys(x for x in row if x not in {"", "N/D", "None"})) or "N/D", axis=1)
    view["Dispositivo"] = view["Dispositivo"].map(simple_device)
    view = view[["Usuario", "Evento", "Correcto", "Fecha y hora", "Ubicación", "IP", "Dispositivo"]]
    st.dataframe(view, hide_index=True, width="stretch")
    st.download_button("Descargar auditoría CSV", view.to_csv(index=False).encode("utf-8-sig"),
                       "auditoria_snoopy.csv", "text/csv")


register_page_open()

if "user" not in st.session_state:
    login_screen()
    st.stop()

user = st.session_state.user
if user.get("must_change_password"):
    password_page(forced=True)
    st.stop()

with st.sidebar:
    st.markdown("### Snoopy IA")
    st.caption(f"Extractor Bancario IA · {AUTHOR}")
    st.caption(f"{user['full_name']} · {user['role']}")
    age_days = password_age_days(user.get("password_changed_at"))
    if age_days is not None and age_days >= 90:
        st.warning(f"Tu clave tiene {age_days} días. Se recomienda actualizarla.")
    pages = ["Extractor"]
    if user["role"] == "Administrador":
        pages += ["Gestión", "Usuarios", "Auditoría"]
    pages += ["Mi clave"]
    page = st.radio("Navegación", pages)
    st.divider()
    st.caption(f"Último acceso: {datetime_ar(user.get('last_login'))} · Argentina")
    if st.button("Cerrar sesión", width="stretch"):
        log_access(user["username"], True, "LOGOUT")
        st.session_state.clear()
        st.rerun()

if page == "Extractor":
    extractor_page()
elif page == "Mi clave":
    password_page()
elif page == "Gestión":
    management_page()
elif page == "Usuarios":
    admin_users_page()
else:
    audit_page()
